"""
Cell Border - Extract border styles from Excel cells and convert to draw.io shapes
"""

from typing import Dict, List, Tuple, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


class CellBorder:
    """Represents a border extracted from an Excel cell"""

    def __init__(
        self,
        cell_address: str,
        border_type: str,  # "top", "bottom", "left", "right"
        style: str,
        color: str,
        x: float,
        y: float,
        width: float,
        height: float,
    ):
        self.cell_address = cell_address
        self.border_type = border_type
        self.style = style  # solid, dashed, dotted, double, medium, thick, thin, none
        self.color = color
        self.x = x
        self.y = y
        self.width = width
        self.height = height

    def to_drawio_shape(self) -> dict:
        """Convert to draw.io shape (line) representation"""
        # Define start and end points based on border position
        if self.border_type == "top":
            x1, y1 = self.x, self.y
            x2, y2 = self.x + self.width, self.y
        elif self.border_type == "bottom":
            x1, y1 = self.x, self.y + self.height
            x2, y2 = self.x + self.width, self.y + self.height
        elif self.border_type == "left":
            x1, y1 = self.x, self.y
            x2, y2 = self.x, self.y + self.height
        elif self.border_type == "right":
            x1, y1 = self.x + self.width, self.y
            x2, y2 = self.x + self.width, self.y + self.height
        else:
            x1, y1 = self.x, self.y
            x2, y2 = self.x + self.width, self.y + self.height

        # Map style to draw.io line style
        style_map = {
            "solid": "solid",
            "dashed": "dashed",
            "dotted": "dotted",
            "double": "double",
            "medium": "solid",
            "thick": "solid",
            "thin": "solid",
            "none": "none",
        }
        dash_pattern = ""
        if self.style == "dashed":
            dash_pattern = "dashPattern=5 5;"
        elif self.style == "dotted":
            dash_pattern = "dashPattern=1 3;"
        elif self.style == "double":
            dash_pattern = "dashed=0;"  # double needs special handling

        # Line width
        width_map = {
            "thin": 1,
            "medium": 2,
            "thick": 3,
            "solid": 1,
            "dashed": 1,
            "dotted": 1,
            "double": 1,
        }
        stroke_width = width_map.get(self.style, 1)

        shape_type = "endArrow=none;startArrow=none;"
        if self.style == "double":
            shape_type = "shape=line;dashed=0;"

        return {
            "type": "line",
            "points": [(x1, y1), (x2, y2)],
            "style": f"endArrow=none;startArrow=none;strokeColor={self.color};strokeWidth={stroke_width};{dash_pattern}",
            "border_type": self.border_type,
        }


class CellGrid:
    """Represents the grid of cells with borders"""

    def __init__(self, worksheet: Worksheet):
        self.worksheet = worksheet
        self.merged_ranges = list(worksheet.merged_cells.ranges)

    def get_cell_position(self, row: int, col: int) -> Tuple[float, float, float, float]:
        """Get cell position and size in EMUs"""
        # Get column width
        col_letter = get_column_letter(col)
        col_dim = self.worksheet.column_dimensions.get(col_letter)
        col_width = col_dim.width if col_dim and col_dim.width else 8.43  # default
        col_width_emu = col_width * 914400 / 8.43  # convert to EMUs

        # Get row height
        row_dim = self.worksheet.row_dimensions.get(row)
        row_height = row_dim.height if row_dim and row_dim.height else 15  # default
        row_height_emu = row_height * 914400 / 72  # convert to EMUs (points)

        # Calculate position
        x = sum(
            (self.worksheet.column_dimensions.get(get_column_letter(c)).width or 8.43)
            if self.worksheet.column_dimensions.get(get_column_letter(c))
            else 8.43
            for c in range(1, col)
        )
        x_emu = x * 914400 / 8.43

        y = sum(
            (self.worksheet.row_dimensions.get(r).height or 15)
            if self.worksheet.row_dimensions.get(r)
            else 15
            for r in range(1, row)
        )
        y_emu = y * 914400 / 72

        return x_emu, y_emu, col_width_emu, row_height_emu

    def is_merged_cell(self, row: int, col: int) -> Optional[Tuple[int, int, int, int]]:
        """Check if cell is part of a merged range, return (min_row, max_row, min_col, max_col)"""
        for merged_range in self.merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return (
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )
        return None

    def get_merged_cell_bounds(
        self, min_row: int, max_row: int, min_col: int, max_col: int
    ) -> Tuple[float, float, float, float]:
        """Calculate the bounds of a merged cell range"""
        x, y = 0, 0
        width, height = 0, 0

        for col in range(min_col, max_col + 1):
            col_letter = get_column_letter(col)
            col_dim = self.worksheet.column_dimensions.get(col_letter)
            col_width = col_dim.width if col_dim and col_dim.width else 8.43
            width += col_width

        for row in range(min_row, max_row + 1):
            row_dim = self.worksheet.row_dimensions.get(row)
            row_height = row_dim.height if row_dim and row_dim.height else 15
            height += row_height

        # Calculate x, y position
        for col in range(1, min_col):
            col_letter = get_column_letter(col)
            col_dim = self.worksheet.column_dimensions.get(col_letter)
            col_width = col_dim.width if col_dim and col_dim.width else 8.43
            x += col_width

        for row in range(1, min_row):
            row_dim = self.worksheet.row_dimensions.get(row)
            row_height = row_dim.height if row_dim and row_dim.height else 15
            y += row_height

        # Convert to EMUs
        x_emu = x * 914400 / 8.43
        y_emu = y * 914400 / 72
        width_emu = width * 914400 / 8.43
        height_emu = height * 914400 / 72

        return x_emu, y_emu, width_emu, height_emu


def extract_borders(worksheet: Worksheet, max_rows: int = 100, max_cols: int = 50) -> List[CellBorder]:
    """Extract all borders from a worksheet"""
    borders = []
    grid = CellGrid(worksheet)

    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            cell = worksheet.cell(row=row, column=col)

            # Skip cells that are part of merged ranges (except top-left)
            merged = grid.is_merged_cell(row, col)
            if merged:
                min_row, max_row, min_col, max_col = merged
                if row != min_row or col != min_col:
                    continue  # Skip non-top-left cells of merged ranges

            # Get cell position
            if merged:
                x, y, width, height = grid.get_merged_cell_bounds(
                    min_row, max_row, min_col, max_col
                )
            else:
                x, y, width, height = grid.get_cell_position(row, col)

            # Get border
            if cell.border and cell.border.top:
                border = cell.border.top
                if border.style and border.style != "none":
                    color = _get_color(border.color)
                    borders.append(
                        CellBorder(
                            cell_address=f"{get_column_letter(col)}{row}",
                            border_type="top",
                            style=border.style,
                            color=color,
                            x=x,
                            y=y,
                            width=width,
                            height=height,
                        )
                    )

            if cell.border and cell.border.bottom:
                border = cell.border.bottom
                if border.style and border.style != "none":
                    color = _get_color(border.color)
                    borders.append(
                        CellBorder(
                            cell_address=f"{get_column_letter(col)}{row}",
                            border_type="bottom",
                            style=border.style,
                            color=color,
                            x=x,
                            y=y,
                            width=width,
                            height=height,
                        )
                    )

            if cell.border and cell.border.left:
                border = cell.border.left
                if border.style and border.style != "none":
                    color = _get_color(border.color)
                    borders.append(
                        CellBorder(
                            cell_address=f"{get_column_letter(col)}{row}",
                            border_type="left",
                            style=border.style,
                            color=color,
                            x=x,
                            y=y,
                            width=width,
                            height=height,
                        )
                    )

            if cell.border and cell.border.right:
                border = cell.border.right
                if border.style and border.style != "none":
                    color = _get_color(border.color)
                    borders.append(
                        CellBorder(
                            cell_address=f"{get_column_letter(col)}{row}",
                            border_type="right",
                            style=border.style,
                            color=color,
                            x=x,
                            y=y,
                            width=width,
                            height=height,
                        )
                    )

    return borders


def _get_color(color) -> str:
    """Extract color from Color object"""
    if not color:
        return "#000000"

    if hasattr(color, "rgb") and color.rgb:
        rgb = color.rgb
        if len(rgb) == 8:
            rgb = rgb[2:]  # Remove alpha
        return f"#{rgb.upper()}"
    elif hasattr(color, "theme"):
        # Theme colors - return default
        return "#000000"
    elif hasattr(color, "indexed"):
        # Indexed colors - return default
        return "#000000"

    return "#000000"
