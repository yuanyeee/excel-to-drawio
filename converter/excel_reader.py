"""
Excel Reader - Extract shapes from Excel worksheets including cells and borders
Enhanced with deep group nesting support, custom geometry, and VML parsing
"""

import zipfile
import re
from typing import Optional, List, Dict, Tuple, Any
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cell_border import extract_borders, CellGrid

# Namespaces used in DrawingML
DML_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

# Register namespaces to avoid ns0, ns1 prefixes
for prefix, uri in DML_NS.items():
    ET.register_namespace(prefix, uri)
ET.register_namespace("v", "urn:schemas-microsoft-com:vml")
ET.register_namespace("o", "urn:schemas-microsoft-com:office:office")
ET.register_namespace("x", "urn:schemas-microsoft-com:office:excel")
ET.register_namespace("pvml", "urn:schemas-microsoft-com:office:powerpoint")
ET.register_namespace("p", "http://schemas.openxmlformats.org/presentationml/2006/main")


class Shape:
    """Represents a shape extracted from Excel"""

    def __init__(
        self,
        shape_id: int,
        name: str,
        type: str,
        x: float,
        y: float,
        width: float,
        height: float,
        text: str = "",
        style: dict = None,
        source: str = "shape",  # "shape" or "cell" or "dml" or "vml"
        geometry: str = None,  # For custom geometry: "rect", "path", etc.
        path_data: str = None,  # SVG path data for custom geometry
    ):
        self.id = shape_id
        self.name = name
        self.type = type
        self.x = x  # EMUs
        self.y = y  # EMUs
        self.width = width  # EMUs
        self.height = height  # EMUs
        self.text = text
        self.style = style or {}
        self.source = source
        self.geometry = geometry  # "rect", "path", etc.
        self.path_data = path_data  # SVG path for custom shapes

    def to_pixels(self):
        """Convert EMUs to pixels (96 DPI)"""
        return {
            "x": self.x / 914400 * 96,
            "y": self.y / 914400 * 96,
            "width": self.width / 914400 * 96,
            "height": self.height / 914400 * 96,
        }


class Connector:
    """Represents a connector/arrow from Excel"""

    def __init__(
        self,
        shape_id: int,
        name: str,
        type: str,
        points: list = None,
        style: dict = None,
    ):
        self.id = shape_id
        self.name = name
        self.type = type
        self.points = points or []  # List of (x, y) tuples in EMUs
        self.style = style or {}


class ExcelReader:
    """Read Excel files and extract shapes and cell-based diagrams"""
    SKIP_FILL_COLORS = {
        "FFFFFF", "FFFFFE", "F2F2F2", "F3F3F3", "EBEBEB", "E7E6E6", "EEECE1",
        "D9D9D9", "BFBFBF", "000000", "0D0D0D",
    }

    def __init__(self, filepath: str, sheet_names: list = None, include_cells: bool = True):
        self.filepath = filepath
        self.sheet_names = sheet_names
        self.include_cells = include_cells
        self.wb = load_workbook(filepath, data_only=True)
        self._zip_path = filepath  # For accessing raw XML

    def read_all(self):
        """Read all sheets and return their shapes"""
        sheets_data = {}

        target_sheets = (
            self.sheet_names
            if self.sheet_names
            else self.wb.sheetnames
        )

        for sheet_name in target_sheets:
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                shapes, connectors = self._extract_shapes(ws)
                sheets_data[sheet_name] = {
                    "shapes": shapes,
                    "connectors": connectors,
                    "title": sheet_name,
                }

        return sheets_data

    def _extract_shapes(self, worksheet):
        """Extract shapes and connectors from a worksheet"""
        shapes = []
        connectors = []
        shape_id_counter = 1

        # 1. Extract from DrawingML (drawing*.xml) with deep group support
        dml_shapes, dml_connectors = self._extract_dml_shapes(worksheet, shape_id_counter)
        shapes.extend(dml_shapes)
        connectors.extend(dml_connectors)
        shape_id_counter += len(dml_shapes)

        # 2. Extract from VML drawings (vmlDrawing*.vml)
        vml_shapes = self._extract_vml_shapes(worksheet, shape_id_counter)
        shapes.extend(vml_shapes)
        shape_id_counter += len(vml_shapes)

        # 3. Extract from openpyxl's native shape parsing (fallback)
        openpyxl_shapes, openpyxl_connectors = self._extract_drawing_shapes(worksheet, shape_id_counter)
        for s in openpyxl_shapes:
            # Avoid duplicates by checking if already extracted via DML
            if not self._is_duplicate_shape(s, shapes):
                shapes.append(s)
        connectors.extend(openpyxl_connectors)
        shape_id_counter += len(openpyxl_shapes)

        # 4. Extract cell-based shapes (merged cells with content)
        if self.include_cells:
            cell_shapes = self._extract_cell_shapes(worksheet, shape_id_counter)
            shapes.extend(cell_shapes)

        return shapes, connectors

    def _is_duplicate_shape(self, shape: Shape, shapes: List[Shape]) -> bool:
        """Check if shape is likely a duplicate"""
        for s in shapes:
            if abs(s.x - shape.x) < 100 and abs(s.y - shape.y) < 100 and s.name == shape.name:
                return True
        return False

    def _extract_dml_shapes(self, worksheet: Worksheet, start_id: int = 1) -> Tuple[List[Shape], List[Connector]]:
        """
        Extract shapes from DrawingML (xdr:) XML with deep group nesting support.
        This handles the complex nested group structures in the network flow diagrams.
        """
        shapes = []
        connectors = []
        shape_id = start_id

        try:
            # Find the drawing IDs for this worksheet
            drawing_ids = self._get_worksheet_drawing_ids(worksheet)
            if not drawing_ids:
                return shapes, connectors

            # Extract EMU constants
            # 1 inch = 914400 EMU, 1 pixel = 914400/96 = 9525 EMU

            # Read the DrawingML XML directly from the zip
            with zipfile.ZipFile(self._zip_path, "r") as zf:
                for drawing_id in drawing_ids:
                    drawing_path = f"xl/drawings/drawing{drawing_id}.xml"
                    if drawing_path not in zf.namelist():
                        continue

                    try:
                        content = zf.read(drawing_path)
                        root = ET.fromstring(content)

                        # Parse anchor-based drawing
                        for anchor in root.iter():
                            if anchor.tag.endswith("}oneCellAnchor") or anchor.tag == "oneCellAnchor":
                                anchor_shapes, anchor_connectors = self._parse_one_cell_anchor(
                                    anchor, 0, 0, [], shape_id
                                )
                                shapes.extend(anchor_shapes)
                                connectors.extend(anchor_connectors)
                                shape_id += len(anchor_shapes)
                            elif anchor.tag.endswith("}twoCellAnchor") or anchor.tag == "twoCellAnchor":
                                from_elem = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
                                if from_elem is None:
                                    from_elem = anchor.find("from")
                                col = 0
                                col_off = 0
                                row = 0
                                row_off = 0
                                if from_elem is not None:
                                    col_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
                                    col_off_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}colOff")
                                    row_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
                                    row_off_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}rowOff")
                                    if col_elem is not None:
                                        try: col = int(col_elem.text)
                                        except: pass
                                    if col_off_elem is not None:
                                        try: col_off = int(col_off_elem.text)
                                        except: pass
                                    if row_elem is not None:
                                        try: row = int(row_elem.text)
                                        except: pass
                                    if row_off_elem is not None:
                                        try: row_off = int(row_off_elem.text)
                                        except: pass

                                anchor_shapes, anchor_connectors = self._parse_one_cell_anchor(
                                    anchor, col, col_off, [row, row_off], shape_id
                                )
                                shapes.extend(anchor_shapes)
                                connectors.extend(anchor_connectors)
                                shape_id += len(anchor_shapes)

                    except Exception as e:
                        continue

        except Exception as e:
            pass

        return shapes, connectors

    def _parse_one_cell_anchor(self, anchor, anchor_col, anchor_col_off, anchor_row_info, start_id):
        """
        Parse a drawing anchor (oneCellAnchor or twoCellAnchor).
        Handles deeply nested group structures.
        """
        shapes = []
        connectors = []
        shape_id = start_id

        # Get anchor offset
        from_elem = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
        if from_elem is None:
            from_elem = anchor.find("from")

        anchor_x = 0
        anchor_y = 0
        if from_elem is not None:
            col_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
            col_off_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}colOff")
            row_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
            row_off_elem = from_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}rowOff")

            if col_elem is not None and col_elem.text:
                try:
                    col = int(col_elem.text)
                    col_off = int(col_off_elem.text) if col_off_elem is not None and col_off_elem.text else 0
                    anchor_x = col * 922880 + col_off  # Standard column width + offset
                except ValueError:
                    anchor_x = 0

            if row_elem is not None and row_elem.text:
                try:
                    row = int(row_elem.text)
                    row_off = int(row_off_elem.text) if row_off_elem is not None and row_off_elem.text else 0
                    anchor_y = row * 255780 + row_off  # Standard row height + offset
                except ValueError:
                    anchor_y = 0

        # Get extent from anchor (for absolute positioning)
        ext_elem = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}ext")
        if ext_elem is None:
            ext_elem = anchor.find("ext")

        anchor_width = 0
        anchor_height = 0
        if ext_elem is not None:
            cx = ext_elem.get("cx", "0")
            cy = ext_elem.get("cy", "0")
            try:
                anchor_width = int(cx)
                anchor_height = int(cy)
            except ValueError:
                pass

        # Walk children of anchor looking for shapes, groups, connectors
        # The immediate child can be xdr:sp, xdr:grpSp, or xdr:cxnSp
        for child in anchor:
            local_name = child.tag.split("}")[-1] if "}" in child.tag else child.tag

            if local_name == "grpSp":
                # Process group - recursively walk to find actual shapes
                group_shapes, group_connectors, shape_id = self._parse_group_element(
                    child, anchor_x, anchor_y, [], shape_id
                )
                shapes.extend(group_shapes)
                connectors.extend(group_connectors)
            elif local_name == "sp":
                # Direct shape in anchor
                shape = self._parse_sp_element(child, anchor_x, anchor_y, [], shape_id)
                if shape:
                    shapes.append(shape)
                    shape_id += 1
            elif local_name == "cxnSp":
                # Connector
                connector = self._parse_cxn_element(child, anchor_x, anchor_y, [], shape_id)
                if connector:
                    connectors.append(connector)
                    shape_id += 1

        return shapes, connectors

    def _parse_group_element(self, grp_elem, parent_x, parent_y, parent_transform, shape_id):
        """
        Recursively parse a group element (xdr:grpSp) and all nested children.
        Properly calculates absolute coordinates by accumulating transforms.
        """
        shapes = []
        connectors = []

        # Get group transform info
        grp_sp_pr = grp_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}grpSpPr")
        if grp_sp_pr is None:
            grp_sp_pr = grp_elem.find("grpSpPr")

        xfrm_elem = None
        if grp_sp_pr is not None:
            xfrm_elem = grp_sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}xfrm")
            if xfrm_elem is None:
                xfrm_elem = grp_sp_pr.find("xfrm")

        # Parse group transform: off, ext, chOff, chExt
        grp_off_x = 0
        grp_off_y = 0
        grp_ext_cx = 0
        grp_ext_cy = 0
        grp_ch_off_x = 0
        grp_ch_off_y = 0
        grp_ch_ext_cx = 0
        grp_ch_ext_cy = 0

        if xfrm_elem is not None:
            off = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}off")
            ext = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}ext")
            ch_off = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}chOff")
            ch_ext = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}chExt")

            if off is not None:
                grp_off_x = int(off.get("x", 0))
                grp_off_y = int(off.get("y", 0))
            if ext is not None:
                grp_ext_cx = int(ext.get("cx", 0))
                grp_ext_cy = int(ext.get("cy", 0))
            if ch_off is not None:
                grp_ch_off_x = int(ch_off.get("x", 0))
                grp_ch_off_y = int(ch_off.get("y", 0))
            if ch_ext is not None:
                grp_ch_ext_cx = int(ch_ext.get("cx", 0))
                grp_ch_ext_cy = int(ch_ext.get("cy", 0))

        # Calculate scale factors for child coordinates
        scale_x = grp_ext_cx / grp_ch_ext_cx if grp_ch_ext_cx != 0 else 1
        scale_y = grp_ext_cy / grp_ch_ext_cy if grp_ch_ext_cy != 0 else 1

        # Build the cumulative transform for this group
        current_transform = {
            "offset_x": parent_x + grp_off_x,
            "offset_y": parent_y + grp_off_y,
            "scale_x": parent_transform[-1]["scale_x"] * scale_x if parent_transform else scale_x,
            "scale_y": parent_transform[-1]["scale_y"] * scale_y if parent_transform else scale_y,
            "ch_off_x": grp_ch_off_x,
            "ch_off_y": grp_ch_off_y,
        }

        # Recurse into children
        for child in grp_elem:
            local_name = child.tag.split("}")[-1] if "}" in child.tag else child.tag

            if local_name == "grpSp":
                # Nested group
                nested_shapes, nested_connectors, shape_id = self._parse_group_element(
                    child, parent_x + grp_off_x, parent_y + grp_off_y,
                    parent_transform + [current_transform], shape_id
                )
                shapes.extend(nested_shapes)
                connectors.extend(nested_connectors)
            elif local_name == "sp":
                # Shape in group
                shape = self._parse_sp_element(
                    child, parent_x + grp_off_x, parent_y + grp_off_y,
                    parent_transform + [current_transform], shape_id
                )
                if shape:
                    shapes.append(shape)
                    shape_id += 1
            elif local_name == "cxnSp":
                # Connector in group
                connector = self._parse_cxn_element(
                    child, parent_x + grp_off_x, parent_y + grp_off_y,
                    parent_transform + [current_transform], shape_id
                )
                if connector:
                    connectors.append(connector)
                    shape_id += 1

        return shapes, connectors, shape_id

    def _parse_sp_element(self, sp_elem, parent_x, parent_y, parent_transform, shape_id):
        """
        Parse a shape element (xdr:sp) and compute absolute coordinates.
        """
        # Get name
        nv_sp_pr = sp_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}nvSpPr")
        if nv_sp_pr is None:
            nv_sp_pr = sp_elem.find("nvSpPr")

        name = f"Shape_{shape_id}"
        if nv_sp_pr is not None:
            c_nv_pr = nv_sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}cNvPr")
            if c_nv_pr is None:
                c_nv_pr = nv_sp_pr.find("cNvPr")
            if c_nv_pr is not None:
                name = c_nv_pr.get("name", name)
                id_val = c_nv_pr.get("id")
                if id_val:
                    try:
                        shape_id = int(id_val)
                    except ValueError:
                        pass

        # Get transform
        sp_pr = sp_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}spPr")
        if sp_pr is None:
            sp_pr = sp_elem.find("spPr")

        x = parent_x
        y = parent_y
        width = 95250  # Default 10px in EMU
        height = 95250

        # Get the shape's xfrm
        xfrm_elem = None
        if sp_pr is not None:
            xfrm_elem = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}xfrm")
            if xfrm_elem is None:
                xfrm_elem = sp_pr.find("xfrm")

        if xfrm_elem is not None:
            off = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}off")
            ext = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}ext")

            if off is not None:
                local_x = int(off.get("x", 0))
                local_y = int(off.get("y", 0))

                # Apply scale from parent transform
                if parent_transform:
                    scale_x = parent_transform[-1]["scale_x"]
                    scale_y = parent_transform[-1]["scale_y"]
                    ch_off_x = parent_transform[-1]["ch_off_x"]
                    ch_off_y = parent_transform[-1]["ch_off_y"]

                    # Transform local coordinates through the group hierarchy
                    if ch_off_x != 0 or ch_off_y != 0:
                        local_x = local_x - ch_off_x
                        local_y = local_y - ch_off_y

                    x = parent_x + local_x * scale_x
                    y = parent_y + local_y * scale_y
                else:
                    x = parent_x + local_x
                    y = parent_y + local_y

            if ext is not None:
                local_cx = int(ext.get("cx", 95250))
                local_cy = int(ext.get("cy", 95250))

                if parent_transform:
                    scale_x = parent_transform[-1]["scale_x"]
                    scale_y = parent_transform[-1]["scale_y"]
                    width = local_cx * scale_x
                    height = local_cy * scale_y
                else:
                    width = local_cx
                    height = local_cy

        # Get geometry type
        shape_type = "rectangle"
        geometry_type = "rect"
        path_data = None

        if sp_pr is not None:
            # Check for custom geometry (a:custGeom)
            cust_geom = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}custGeom")
            if cust_geom is None:
                cust_geom = sp_pr.find("custGeom")

            if cust_geom is not None:
                shape_type = "custom"
                geometry_type = "path"
                path_data = self._parse_cust_geom(cust_geom)

            # Otherwise check preset geometry
            if geometry_type == "rect":
                prst_geom = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}prstGeom")
                if prst_geom is None:
                    prst_geom = sp_pr.find("prstGeom")
                if prst_geom is not None:
                    prst = prst_geom.get("prst", "rect")
                    shape_type = self._map_preset_geometry(prst)

        # Check for off-page connector (special handling)
        if "OffpageConnector" in sp_elem.get("spPr", "") or (sp_pr is not None and "flowChartOffpageConnector" in ET.tostring(sp_pr, encoding="unicode")):
            shape_type = "offpageConnector"

        # Get text content
        text = ""
        tx_body = sp_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}txBody")
        if tx_body is None:
            tx_body = sp_elem.find("txBody")

        if tx_body is not None:
            text = self._extract_text_from_txbody(tx_body)

        # Get style (fill, line)
        style = self._extract_sp_style(sp_pr)

        # Handle off-page connector type
        if sp_pr is not None:
            prst_geom = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}prstGeom")
            if prst_geom is None:
                prst_geom = sp_pr.find("prstGeom")
            if prst_geom is not None:
                prst = prst_geom.get("prst", "")
                if prst == "flowChartOffpageConnector":
                    shape_type = "offpageConnector"

        return Shape(
            shape_id=shape_id,
            name=name,
            type=shape_type,
            x=x,
            y=y,
            width=width,
            height=height,
            text=text,
            style=style,
            source="dml",
            geometry=geometry_type,
            path_data=path_data,
        )

    def _parse_cxn_element(self, cxn_elem, parent_x, parent_y, parent_transform, shape_id):
        """
        Parse a connector element (xdr:cxnSp).
        """
        name = f"Connector_{shape_id}"

        nv_cxn_sp_pr = cxn_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}nvCxnSpPr")
        if nv_cxn_sp_pr is None:
            nv_cxn_sp_pr = cxn_elem.find("nvCxnSpPr")

        if nv_cxn_sp_pr is not None:
            c_nv_pr = nv_cxn_sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}cNvPr")
            if c_nv_pr is None:
                c_nv_pr = nv_cxn_sp_pr.find("cNvPr")
            if c_nv_pr is not None:
                name = c_nv_pr.get("name", name)

        sp_pr = cxn_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}spPr")
        if sp_pr is None:
            sp_pr = cxn_elem.find("spPr")

        # Get geometry
        connector_type = "straightConnector1"
        if sp_pr is not None:
            prst_geom = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}prstGeom")
            if prst_geom is None:
                prst_geom = sp_pr.find("prstGeom")
            if prst_geom is not None:
                connector_type = prst_geom.get("prst", "straightConnector1")

        # Get points
        xfrm_elem = None
        if sp_pr is not None:
            xfrm_elem = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}xfrm")
            if xfrm_elem is None:
                xfrm_elem = sp_pr.find("xfrm")

        points = []
        if xfrm_elem is not None:
            off = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}off")
            ext = xfrm_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}ext")

            if off is not None and ext is not None:
                local_x = int(off.get("x", 0))
                local_y = int(off.get("y", 0))
                cx = int(ext.get("cx", 0))
                cy = int(ext.get("cy", 0))

                if parent_transform:
                    scale_x = parent_transform[-1]["scale_x"]
                    scale_y = parent_transform[-1]["scale_y"]
                    ch_off_x = parent_transform[-1]["ch_off_x"]
                    ch_off_y = parent_transform[-1]["ch_off_y"]

                    if ch_off_x != 0 or ch_off_y != 0:
                        local_x = local_x - ch_off_x
                        local_y = local_y - ch_off_y

                    x = parent_x + local_x * scale_x
                    y = parent_y + local_y * scale_y
                else:
                    x = parent_x + local_x
                    y = parent_y + local_y

                points = [(x, y), (x + cx, y + cy)]

        style = self._extract_sp_style(sp_pr)

        return Connector(
            shape_id=shape_id,
            name=name,
            type=connector_type,
            points=points,
            style=style,
        )

    def _parse_cust_geom(self, cust_geom_elem) -> str:
        """
        Parse custom geometry (a:custGeom) to SVG path data.
        Returns an SVG path string.
        """
        path_cmds = []

        path_lst = cust_geom_elem.find("{http://schemas.openxmlformats.org/drawingml/2006/main}pathLst")
        if path_lst is None:
            path_lst = cust_geom_elem.find("pathLst")

        if path_lst is not None:
            for path_elem in path_lst:
                local_name = path_elem.tag.split("}")[-1]
                if local_name == "path":
                    path_data = path_elem.get("path", "")
                    path_cmds.append(path_data)

        return " ".join(path_cmds)

    def _extract_text_from_txbody(self, tx_body) -> str:
        """Extract text content from a txBody element."""
        text_parts = []

        for elem in tx_body.iter():
            # Collect text from a:t elements
            if elem.tag.endswith("}t"):
                if elem.text:
                    text_parts.append(elem.text)
            # Also collect text directly in r elements
            elif elem.tag.endswith("}r") or elem.tag == "r":
                for t in elem:
                    if t.tag.endswith("}t") and t.text:
                        text_parts.append(t.text)

        return "".join(text_parts)

    def _extract_sp_style(self, sp_pr) -> dict:
        """Extract style properties from a spPr element."""
        style = {}

        if sp_pr is None:
            return style

        # Fill
        solid_fill = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}solidFill")
        if solid_fill is None:
            solid_fill = sp_pr.find("solidFill")

        if solid_fill is not None:
            srgb = solid_fill.find("{http://schemas.openxmlformats.org/drawingml/2006/main}srgbClr")
            if srgb is None:
                srgb = solid_fill.find("srgbClr")
            if srgb is not None:
                style["fillColor"] = "#" + srgb.get("val", "000000")

            # Scheme color
            scheme_clr = solid_fill.find("{http://schemas.openxmlformats.org/drawingml/2006/main}schemeClr")
            if scheme_clr is None:
                scheme_clr = solid_fill.find("schemeClr")
            if scheme_clr is not None and "fillColor" not in style:
                style["fillColor"] = "scheme:" + scheme_clr.get("val", "")

        # Line
        ln = sp_pr.find("{http://schemas.openxmlformats.org/drawingml/2006/main}ln")
        if ln is None:
            ln = sp_pr.find("ln")

        if ln is not None:
            # Line color
            solid_fill = ln.find("{http://schemas.openxmlformats.org/drawingml/2006/main}solidFill")
            if solid_fill is None:
                solid_fill = ln.find("solidFill")
            if solid_fill is not None:
                srgb = solid_fill.find("{http://schemas.openxmlformats.org/drawingml/2006/main}srgbClr")
                if srgb is None:
                    srgb = solid_fill.find("srgbClr")
                if srgb is not None:
                    style["strokeColor"] = "#" + srgb.get("val", "000000")

            # Line width
            width = ln.get("w")
            if width:
                try:
                    style["strokeWidth"] = int(width) / 12700  # Convert to points (1/12700 inch)
                except ValueError:
                    pass

        return style

    def _map_preset_geometry(self, prst: str) -> str:
        """Map DrawingML preset geometry names to our shape types."""
        mapping = {
            "rect": "rectangle",
            "roundRect": "roundRectangle",
            "ellipse": "ellipse",
            "diamond": "diamond",
            "parallelogram": "parallelogram",
            "trapezoid": "trapezoid",
            "hexagon": "hexagon",
            "octagon": "octagon",
            "triangle": "triangle",
            "pentagon": "pentagon",
            "cross": "cross",
            "star": "star",
            "heart": "heart",
            "lightningBolt": "lightning",
            "flowChartOffpageConnector": "offpageConnector",
            "flowChartProcess": "process",
            "flowChartDecision": "decision",
            "flowChartTerminator": "terminator",
        }
        return mapping.get(prst, prst)

    def _find_worksheet_rels(self, worksheet: Worksheet) -> Optional[str]:
        """Find the worksheet's relationship file path."""
        # In openpyxl, we need to access the underlying zip
        return None

    def _get_worksheet_drawing_ids(self, worksheet: Worksheet) -> List[int]:
        """Get drawing IDs associated with a worksheet."""
        drawing_ids = []

        try:
            # Check worksheet for drawing references
            if hasattr(worksheet, "_drawing"):
                drawing = worksheet._drawing
                if drawing and hasattr(drawing, "id"):
                    drawing_ids.append(drawing.id)

            # Try to get from worksheet relationships
            for rel in worksheet._rels:
                if "drawing" in rel.reltype.lower():
                    # Extract drawing number from target
                    match = re.search(r"drawing(\d+)", rel.target)
                    if match:
                        drawing_ids.append(int(match.group(1)))

        except Exception:
            pass

        # Also try to find all drawings in the zip by sheet name
        if not drawing_ids:
            drawing_ids = self._find_drawings_for_sheet(worksheet.title)

        return drawing_ids

    def _find_drawings_for_sheet(self, sheet_name: str) -> List[int]:
        """Find drawing IDs for a worksheet by examining the zip contents."""
        drawing_ids = []

        try:
            # Read workbook.xml to find sheet relationships
            with zipfile.ZipFile(self._zip_path, "r") as zf:
                # Find workbook.xml
                wb_xml = None
                for name in zf.namelist():
                    if name == "xl/workbook.xml":
                        wb_xml = zf.read(name)
                        break

                if wb_xml is None:
                    return drawing_ids

                wb_root = ET.fromstring(wb_xml)
                ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

                # Find sheet by name
                sheet_id = None
                for sheet in wb_root.iter():
                    if sheet.tag.endswith("}sheet") or sheet.tag == "sheet":
                        name = sheet.get("name", "")
                        if name == sheet_name:
                            sheet_id = sheet.get("sheetId")
                            break

                if not sheet_id:
                    return drawing_ids

                # Find the relationship file
                # xl/worksheets/_rels/sheet*.xml.rels
                rels_files = [n for n in zf.namelist() if re.match(r"xl/worksheets/_rels/sheet\d+\.xml\.rels", n)]

                for rels_file in rels_files:
                    rels_content = zf.read(rels_file)
                    rels_root = ET.fromstring(rels_content)

                    for rel in rels_root:
                        if "drawing" in rel.get("Type", "").lower():
                            target = rel.get("Target", "")
                            match = re.search(r"drawing(\d+)", target)
                            if match:
                                drawing_ids.append(int(match.group(1)))

        except Exception as e:
            pass

        return drawing_ids

    def _extract_vml_shapes(self, worksheet: Worksheet, start_id: int = 1) -> List[Shape]:
        """Extract shapes from VML drawings."""
        shapes = []
        shape_id = start_id

        try:
            from .vml_reader import VMLReader

            vml_reader = VMLReader()

            # Find VML drawing IDs
            vml_ids = self._find_vml_drawings_for_sheet(worksheet.title)

            with zipfile.ZipFile(self._zip_path, "r") as zf:
                for vml_id in vml_ids:
                    vml_path = f"xl/drawings/vmlDrawing{vml_id}.vml"
                    if vml_path not in zf.namelist():
                        continue

                    try:
                        content = zf.read(vml_path).decode("utf-8", errors="replace")
                        vml_shapes = vml_reader.read_string(content)

                        for vml_shape in vml_shapes:
                            # Convert VMLShape to Shape
                            shape = Shape(
                                shape_id=shape_id,
                                name=vml_shape.name,
                                type=vml_shape.type,
                                x=vml_shape.left * 9525,  # px to EMU (9525 EMU per pixel)
                                y=vml_shape.top * 9525,
                                width=vml_shape.width * 9525,
                                height=vml_shape.height * 9525,
                                text=vml_shape.text,
                                style={
                                    "fillColor": vml_shape.fill_color,
                                    "strokeColor": vml_shape.stroke_color,
                                } if vml_shape.fill_color or vml_shape.stroke_color else {},
                                source="vml",
                            )
                            shapes.append(shape)
                            shape_id += 1

                    except Exception:
                        continue

        except Exception:
            pass

        return shapes

    def _find_vml_drawings_for_sheet(self, sheet_name: str) -> List[int]:
        """Find VML drawing IDs for a worksheet."""
        vml_ids = []

        try:
            with zipfile.ZipFile(self._zip_path, "r") as zf:
                # Check xl/drawings/_rels for vmlDrawing references
                rels_files = [n for n in zf.namelist() if re.match(r"xl/drawings/_rels/drawing\d+\.xml\.rels", n)]

                for rels_file in rels_files:
                    try:
                        rels_content = zf.read(rels_file)
                        rels_root = ET.fromstring(rels_content)

                        for rel in rels_root:
                            target = rel.get("Target", "")
                            if "vmlDrawing" in target:
                                match = re.search(r"vmlDrawing(\d+)", target)
                                if match:
                                    vml_ids.append(int(match.group(1)))
                    except Exception:
                        continue

        except Exception:
            pass

        return list(set(vml_ids))

    def _extract_drawing_shapes(self, worksheet, start_id: int = 1):
        """Extract shapes from worksheet drawings (openpyxl fallback)"""
        shapes = []
        connectors = []

        try:
            for shape in worksheet._sheets:
                shape_obj = self._parse_shape(shape, start_id, source="shape")
                if shape_obj:
                    shapes.append(shape_obj)

            for drawing in worksheet._drawings:
                from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
                if isinstance(drawing, SpreadsheetDrawing):
                    for shape in drawing.shapes:
                        shape_obj = self._parse_shape(shape, start_id + len(shapes), source="shape")
                        if shape_obj:
                            shapes.append(shape_obj)
        except Exception:
            pass

        return shapes, connectors

    def _extract_cell_shapes(self, worksheet: Worksheet, start_id: int = 1) -> List[Shape]:
        """Extract shapes from cells (merged cells with content or borders)"""
        shapes = []
        grid = CellGrid(worksheet)
        shape_id = start_id
        fill_only_cells: Dict[Tuple[int, int], str] = {}

        # Extract merged cells with content
        for merged_range in worksheet.merged_cells.ranges:
            cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)

            if cell.value:
                x, y, width, height = grid.get_merged_cell_bounds(
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )

                text = str(cell.value) if cell.value else ""
                style = self._extract_cell_style(cell)

                shapes.append(
                    Shape(
                        shape_id=shape_id,
                        name=f"Cell_{get_column_letter(merged_range.min_col)}{merged_range.min_row}",
                        type="rectangle",
                        x=x,
                        y=y,
                        width=width,
                        height=height,
                        text=text,
                        style=style,
                        source="cell",
                    )
                )
                shape_id += 1

        # Extract cells with fill color
        for row in worksheet.iter_rows():
            for cell in row:
                merged = grid.is_merged_cell(cell.row, cell.column)
                if merged:
                    min_row, max_row, min_col, max_col = merged
                    # Merged cells are handled in the dedicated merged-cell loop above.
                    if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                        continue

                has_fill = self._cell_has_meaningful_fill(cell)
                has_border = self._cell_has_visible_border(cell)
                has_text = cell.value is not None and str(cell.value).strip() != ""

                # Avoid creating noisy shapes for plain text-only cells.
                if not (has_fill or has_border):
                    continue

                # Merge fill-only cells later to reduce draw.io object count.
                if has_fill and not has_border and not has_text:
                    fill_color = self._get_cell_fill_hex(cell)
                    if fill_color:
                        fill_only_cells[(cell.row, cell.column)] = fill_color
                    continue

                x, y, width, height = grid.get_cell_position(cell.row, cell.column)
                text = str(cell.value) if has_text else ""
                style = self._extract_cell_style(cell)

                shapes.append(
                    Shape(
                        shape_id=shape_id,
                        name=f"Cell_{get_column_letter(cell.column)}{cell.row}",
                        type="rectangle",
                        x=x,
                        y=y,
                        width=width,
                        height=height,
                        text=text,
                        style=style,
                        source="cell",
                    )
                )
                shape_id += 1

        # Merge adjacent cells with the same fill color into fewer large shapes.
        for region in self._merge_fill_only_cells(fill_only_cells):
            min_row, max_row, min_col, max_col, fill_color = region
            x, y, _, _ = grid.get_cell_position(min_row, min_col)
            width = 0
            for col in range(min_col, max_col + 1):
                _, _, cell_w, _ = grid.get_cell_position(min_row, col)
                width += cell_w
            height = 0
            for row in range(min_row, max_row + 1):
                _, _, _, cell_h = grid.get_cell_position(row, min_col)
                height += cell_h

            shapes.append(
                Shape(
                    shape_id=shape_id,
                    name=f"CellFill_{get_column_letter(min_col)}{min_row}",
                    type="rectangle",
                    x=x,
                    y=y,
                    width=width,
                    height=height,
                    text="",
                    style={"fillColor": fill_color},
                    source="cell",
                )
            )
            shape_id += 1

        return shapes

    def _merge_fill_only_cells(
        self, fill_cells: Dict[Tuple[int, int], str]
    ) -> List[Tuple[int, int, int, int, str]]:
        """
        Merge adjacent fill-only cells with same color into rectangular regions.
        Returns tuples: (min_row, max_row, min_col, max_col, color).
        """
        merged_regions: List[Tuple[int, int, int, int, str]] = []
        processed: set = set()

        for row, col in sorted(fill_cells.keys()):
            if (row, col) in processed:
                continue
            color = fill_cells[(row, col)]

            max_col = col
            while (
                (row, max_col + 1) in fill_cells
                and fill_cells[(row, max_col + 1)] == color
                and (row, max_col + 1) not in processed
            ):
                max_col += 1

            max_row = row
            while True:
                next_row = max_row + 1
                all_match = True
                for candidate_col in range(col, max_col + 1):
                    key = (next_row, candidate_col)
                    if (
                        key not in fill_cells
                        or fill_cells[key] != color
                        or key in processed
                    ):
                        all_match = False
                        break
                if not all_match:
                    break
                max_row = next_row

            for r in range(row, max_row + 1):
                for c in range(col, max_col + 1):
                    processed.add((r, c))

            merged_regions.append((row, max_row, col, max_col, color))

        return merged_regions

    def _parse_shape(self, shape, shape_id: int, source: str = "shape") -> Optional[Shape]:
        """Parse a single shape object from openpyxl"""
        try:
            name = shape.name if hasattr(shape, "name") else ""
            type_ = shape.type if hasattr(shape, "type") else "rectangle"

            x = getattr(shape, "x", 0) or 0
            y = getattr(shape, "y", 0) or 0
            width = getattr(shape, "width", 0) or 0
            height = getattr(shape, "height", 0) or 0

            text = ""
            if hasattr(shape, "text") and shape.text:
                text = str(shape.text)
            elif hasattr(shape, "value") and shape.value:
                text = str(shape.value)

            style = self._extract_style(shape)

            return Shape(
                shape_id=shape_id,
                name=name,
                type=type_,
                x=x,
                y=y,
                width=width,
                height=height,
                text=text,
                style=style,
                source=source,
            )
        except Exception:
            return None

    def _extract_style(self, shape) -> dict:
        """Extract style properties from a shape"""
        style = {}

        try:
            if hasattr(shape, "fill") and shape.fill:
                fill = shape.fill
                if hasattr(fill, "fgColor") and fill.fgColor:
                    color = fill.fgColor
                    if hasattr(color, "rgb"):
                        style["fillColor"] = self._rgb_to_hex(color.rgb)
                    elif hasattr(color, "theme"):
                        style["fillColor"] = f"theme:{color.theme}"

            if hasattr(shape, "line") and shape.line:
                line = shape.line
                if hasattr(line, "color") and line.color:
                    color = line.color
                    if hasattr(color, "rgb"):
                        style["strokeColor"] = self._rgb_to_hex(color.rgb)
                if hasattr(line, "width"):
                    style["strokeWidth"] = line.width

        except Exception:
            pass

        return style

    def _extract_cell_style(self, cell) -> dict:
        """Extract style properties from a cell"""
        style = {}

        try:
            fill_color = self._get_cell_fill_hex(cell)
            if fill_color:
                style["fillColor"] = fill_color

            if cell.font:
                font = cell.font
                if hasattr(font, "size") and font.size:
                    style["fontSize"] = int(font.size)
                if hasattr(font, "bold") and font.bold:
                    style["fontStyle"] = "bold"
                if hasattr(font, "color") and font.color:
                    color = font.color
                    if hasattr(color, "rgb") and color.rgb:
                        rgb = self._normalize_rgb_value(color.rgb)
                        if rgb:
                            style["fontColor"] = f"#{rgb}"

            if cell.alignment:
                align = cell.alignment
                if hasattr(align, "horizontal") and align.horizontal:
                    style["align"] = align.horizontal
                if hasattr(align, "vertical") and align.vertical:
                    style["verticalAlign"] = align.vertical

        except Exception:
            pass

        return style

    def _cell_has_visible_border(self, cell) -> bool:
        """Return True when a cell has at least one visible border side."""
        border = getattr(cell, "border", None)
        if not border:
            return False
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name, None)
            if side and side.style and side.style != "none":
                return True
        return False

    def _cell_has_meaningful_fill(self, cell) -> bool:
        """Return True when a cell has a non-default solid fill."""
        return self._get_cell_fill_hex(cell) is not None

    def _get_cell_fill_hex(self, cell) -> Optional[str]:
        """Extract meaningful solid fill color as #RRGGBB, otherwise None."""
        if not cell.fill or cell.fill.fill_type != "solid":
            return None
        fg_color = getattr(cell.fill, "fgColor", None)
        if not fg_color:
            return None
        rgb = self._normalize_rgb_value(getattr(fg_color, "rgb", None))
        if not rgb:
            return None
        if rgb in self.SKIP_FILL_COLORS:
            return None
        return f"#{rgb}"

    def _rgb_to_hex(self, rgb: str) -> str:
        """Convert RGB string to hex color"""
        normalized = self._normalize_rgb_value(rgb)
        if not normalized:
            return "#000000"
        return f"#{normalized}"

    def _normalize_rgb_value(self, rgb) -> Optional[str]:
        """Normalize OpenPyXL RGB-like values to 6-char uppercase hex."""
        if rgb is None:
            return None

        # Fast-path for normal string values.
        if isinstance(rgb, str):
            pass
        else:
            # openpyxl may return custom RGB objects that expose `.value`
            try:
                rgb = object.__getattribute__(rgb, "value")
            except Exception:
                return None
            if not isinstance(rgb, str):
                return None

        rgb = rgb.strip()
        if not rgb:
            return None
        if rgb.startswith("#"):
            rgb = rgb[1:]
        rgb = rgb.upper()

        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) != 6:
            return None
        if not re.fullmatch(r"[0-9A-F]{6}", rgb):
            return None
        return rgb

    def close(self):
        """Close the workbook"""
        self.wb.close()
