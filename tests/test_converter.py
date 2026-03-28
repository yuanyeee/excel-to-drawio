"""
Tests for Excel to draw.io Converter
"""

import pytest
import os
import py_compile
import tempfile
from pathlib import Path
import importlib

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from converter.excel_reader import ExcelReader, Shape, Connector
from converter.shape_mapper import ShapeMapper
from converter.drawio_writer import SimpleDrawioWriter, DrawioWriter
from converter.excel_to_drawio import convert_excel_to_drawio


class TestShapeMapper:
    """Test shape type mapping"""

    def test_map_basic_shapes(self):
        mapper = ShapeMapper()

        assert mapper.map_type("rectangle") == "rectangle"
        assert mapper.map_type("ellipse") == "ellipse"
        assert mapper.map_type("diamond") == "diamond"
        assert mapper.map_type("roundRectangle") == "roundRect"
        assert mapper.map_type("textBox") == "text"

    def test_map_unknown_returns_rectangle(self):
        mapper = ShapeMapper()
        assert mapper.map_type("unknown_shape") == "rectangle"
        assert mapper.map_type("") == "rectangle"
        assert mapper.map_type(None) == "rectangle"

    def test_build_style_string(self):
        mapper = ShapeMapper()
        style = {
            "shape": "rectangle",
            "fillColor": "#4A90D9",
            "strokeColor": "#2E5C8A",
            "strokeWidth": 2,
        }
        result = mapper.build_style_string(style)

        assert "fillColor=#4A90D9" in result
        assert "strokeColor=#2E5C8A" in result
        assert "strokeWidth=2" in result
        assert "whiteSpace=wrap" in result
        assert "html=1" in result


class TestShape:
    """Test Shape class"""

    def test_shape_creation(self):
        shape = Shape(
            shape_id=1,
            name="Test Shape",
            type="rectangle",
            x=914400,  # 1 inch in EMUs
            y=914400,
            width=1828800,  # 2 inches
            height=914400,
            text="Hello",
        )

        assert shape.id == 1
        assert shape.name == "Test Shape"
        assert shape.type == "rectangle"
        assert shape.text == "Hello"

    def test_shape_to_pixels(self):
        shape = Shape(
            shape_id=1,
            name="Test",
            type="rectangle",
            x=914400,
            y=914400,
            width=1828800,
            height=914400,
        )

        pixels = shape.to_pixels()
        assert pixels["x"] == 96  # 1 inch at 96 DPI
        assert pixels["width"] == 192  # 2 inches at 96 DPI


class TestSimpleDrawioWriter:
    """Test draw.io writer"""

    def test_write_simple_drawio(self):
        shapes = [
            Shape(
                shape_id=1,
                name="Box1",
                type="rectangle",
                x=914400,
                y=914400,
                width=1828800,
                height=914400,
                text="Box 1",
            ),
            Shape(
                shape_id=2,
                name="Box2",
                type="ellipse",
                x=3657600,
                y=914400,
                width=1828800,
                height=914400,
                text="Box 2",
            ),
        ]

        with tempfile.NamedTemporaryFile(suffix=".drawio", delete=False) as f:
            temp_path = f.name

        try:
            writer = SimpleDrawioWriter(shapes, title="Test Sheet")
            writer.write(temp_path)

            assert os.path.exists(temp_path)
            with open(temp_path, "r", encoding="utf-8") as f:
                content = f.read()

            assert "<mxfile" in content
            assert "<diagram" in content
            assert "Box 1" in content
            assert "Box 2" in content
            assert "</mxfile>" in content
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)


class TestDrawioWriter:
    """Test multi-page draw.io writer"""

    def test_connector_geometry_uses_source_target_points(self):
        connector = Connector(
            shape_id=10,
            name="Arrow",
            type="straightConnector1",
            points=[(0, 0), (914400, 914400)],
            style={"endArrow": "block"},
        )

        sheets_data = {
            "Sheet1": {
                "shapes": [],
                "connectors": [connector],
                "title": "Sheet1",
            }
        }

        with tempfile.NamedTemporaryFile(suffix=".drawio", delete=False) as f:
            temp_path = f.name

        try:
            writer = DrawioWriter(sheets_data)
            writer.write(temp_path)

            with open(temp_path, "r", encoding="utf-8") as f:
                content = f.read()

            assert 'mxPoint x="0.0" y="0.0" as="sourcePoint"' in content
            assert 'mxPoint x="96.0" y="96.0" as="targetPoint"' in content
            assert "<Array>" not in content
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)


class TestHighLevelConverter:
    def test_convert_excel_to_drawio_creates_single_file(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1["A1"] = "Hello"
        ws2 = wb.create_sheet("S2")
        ws2["A1"] = "World"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f_in:
            input_path = f_in.name
        with tempfile.NamedTemporaryFile(suffix=".drawio", delete=False) as f_out:
            output_path = f_out.name

        try:
            wb.save(input_path)
            result = convert_excel_to_drawio(input_path, output_path)

            assert os.path.exists(output_path)
            assert result.sheet_names == ["S1", "S2"]
            with open(output_path, "r", encoding="utf-8") as f:
                content = f.read()
            assert "name=\"S1\"" in content
            assert "name=\"S2\"" in content
        finally:
            if os.path.exists(input_path):
                os.unlink(input_path)
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestExcelReader:
    """Test Excel reader (requires actual Excel file)"""

    def test_reader_initialization(self):
        # Just test that we can initialize with a non-existent file
        # Actual file reading would require a test Excel file
        reader = ExcelReader("nonexistent.xlsx")
        assert reader.filepath == "nonexistent.xlsx"

    def test_extract_cell_shapes_skips_plain_text_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "plain text only"
        ws["B1"] = "filled"
        ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FF00FF00")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            reader = ExcelReader(temp_path, include_cells=True)
            shapes, _ = reader._extract_shapes(reader.wb.active)

            cell_shapes = [s for s in shapes if s.source == "cell"]
            names = [s.name for s in cell_shapes]

            assert "Cell_A1" not in names
            assert "Cell_B1" in names
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def test_extract_cell_shapes_merges_adjacent_fill_only_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            reader = ExcelReader(temp_path, include_cells=True)
            shapes, _ = reader._extract_shapes(reader.wb.active)

            cell_shapes = [s for s in shapes if s.source == "cell"]

            assert len(cell_shapes) == 1
            assert cell_shapes[0].style.get("fillColor") == "#FF0000"
            assert cell_shapes[0].width > 914400  # Wider than a single default cell
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def test_rgb_normalization_accepts_rgb_objects(self):
        class DummyRGB:
            def __init__(self, value):
                self.value = value

            def __str__(self):
                return self.value

        wb = Workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            reader = ExcelReader(temp_path, include_cells=True)
            assert reader._rgb_to_hex(DummyRGB("FF00FF00")) == "#00FF00"
            assert reader._rgb_to_hex(DummyRGB("00FF00")) == "#00FF00"
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def test_rgb_normalization_ignores_non_string_rgb_objects(self):
        class BrokenRGB:
            @property
            def value(self):
                raise NameError("name 'rgb' is not defined")

        wb = Workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            reader = ExcelReader(temp_path, include_cells=True)
            assert reader._normalize_rgb_value(BrokenRGB()) is None
            assert reader._rgb_to_hex(BrokenRGB()) == "#000000"
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def test_excel_reader_module_compiles(self):
        module_path = Path(__file__).resolve().parents[1] / "converter" / "excel_reader.py"
        py_compile.compile(str(module_path), doraise=True)

    def test_gui_import_path_compiles_with_converter(self):
        # Regression guard: catches syntax/indent issues that break
        # `python gui_tkinter.py` at import time.
        importlib.import_module("converter")
        importlib.import_module("gui_tkinter")

    def test_read_all_continues_when_one_sheet_fails(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "OK"
        wb.create_sheet("BROKEN")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            reader = ExcelReader(temp_path, include_cells=True)

            original_extract_shapes = reader._extract_shapes

            def _fake_extract_shapes(ws):
                if ws.title == "BROKEN":
                    raise TypeError("'NoneType' object is not iterable")
                return original_extract_shapes(ws)

            reader._extract_shapes = _fake_extract_shapes  # type: ignore[method-assign]
            result = reader.read_all()

            assert "OK" in result
            assert "BROKEN" in result
            assert result["BROKEN"]["shapes"] == []
            assert result["BROKEN"]["connectors"] == []
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
